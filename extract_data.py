import openpyxl, json

wb = openpyxl.load_workbook(r"D:\AI分析\税费分析\2026年税费统计分析表_终版7.23_v2.xlsx", data_only=True)
ws_yr = wb['年度总览-2026']
cities=[]; companies=[]; all_raw=[]
for r in range(2, ws_yr.max_row):
    label=ws_yr.cell(r,1).value; name=ws_yr.cell(r,2).value; nature=ws_yr.cell(r,3).value
    inc=float(ws_yr.cell(r,4).value or 0); vat=float(ws_yr.cell(r,5).value or 0)
    surtax=float(ws_yr.cell(r,6).value or 0); cit=float(ws_yr.cell(r,7).value or 0)
    iit=float(ws_yr.cell(r,8).value or 0); stamp=float(ws_yr.cell(r,9).value or 0)
    cultural=float(ws_yr.cell(r,10).value or 0); tt=float(ws_yr.cell(r,15).value or 0)
    vr=float(ws_yr.cell(r,16).value or 0); tr=float(ws_yr.cell(r,18).value or 0)
    if label and '小计' in str(label):
        city=str(label).replace('【','').replace('小计】','').replace('】','')
        cities.append({'name':city,'income':round(inc,2),'vat':round(vat,2),'surtax':round(surtax,2),
                       'cit':round(cit,2),'iit':round(iit,2),'stamp':round(stamp,2),
                       'cultural':round(cultural,2),'total_tax':round(tt,2),
                       'vat_rate':round(vr*100,4),'total_rate':round(tr*100,4)})
    elif name and '小计' not in str(name) and '合计' not in str(name):
        companies.append({'name':name,'city':label or '','nature':nature or '',
                          'income':round(inc,2),'vat':round(vat,2),'surtax':round(surtax,2),
                          'cit':round(cit,2),'iit':round(iit,2),'stamp':round(stamp,2),
                          'cultural':round(cultural,2),'total_tax':round(tt,2),
                          'vat_rate':round(vr*100,4),'total_rate':round(tr*100,4)})
        all_raw.append({'name':name,'city':label or '','nature':nature or '',
                        'income':inc,'vat':vat,'surtax':surtax,'cit':cit,'iit':iit,
                        'stamp':stamp,'cultural':cultural,'total_tax':tt})

lr=ws_yr.max_row
gt={'income':round(float(ws_yr.cell(lr,4).value or 0),2),'vat':round(float(ws_yr.cell(lr,5).value or 0),2),
    'surtax':round(float(ws_yr.cell(lr,6).value or 0),2),'cit':round(float(ws_yr.cell(lr,7).value or 0),2),
    'iit':round(float(ws_yr.cell(lr,8).value or 0),2),'stamp':round(float(ws_yr.cell(lr,9).value or 0),2),
    'cultural':round(float(ws_yr.cell(lr,10).value or 0),2),'tax':round(float(ws_yr.cell(lr,15).value or 0),2),
    'rate':round(float(ws_yr.cell(lr,18).value or 0)*100,4)}

# 月度趋势
ws_mo=wb['累计同比-城市']
months_order=['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
monthly={}
for r in range(2, ws_mo.max_row+1):
    m=ws_mo.cell(r,1).value; cy=ws_mo.cell(r,2).value
    if cy=='【全集团合计】':
        monthly[m]={'inc26':round(float(ws_mo.cell(r,3).value or 0),2),
                     'inc25':round(float(ws_mo.cell(r,4).value or 0),2),
                     'vat26':round(float(ws_mo.cell(r,6).value or 0),2)}

# 小规模纳税人 + 滚动12月预警
wb25=openpyxl.load_workbook(r"D:\AI分析\税费分析\税费统计表2025 .xlsx", data_only=True)
ws25=wb25['25-汇总']
inc_2025={}
for r in range(4, ws25.max_row+1):
    n=ws25.cell(r,4).value; cy=ws25.cell(r,2).value
    if not n: continue
    ns=str(n).strip()
    if '小计' in ns or '合计' in ns or 'check' in ns: continue
    if ns in ['上海','北京','南京','广州','成都','杭州','武汉','深圳']: continue
    if cy is None or str(cy).strip()=='': continue
    inc_2025[ns]=float(ws25.cell(r,16).value or 0)

small_raw=[c for c in all_raw if '小规模' in str(c['nature'])]
general_raw=[c for c in all_raw if '一般' in str(c['nature'])]
small_detail=[]
for c in small_raw:
    inc_2025_full=inc_2025.get(c['name'],0)
    rolling_12m=inc_2025_full/2 + c['income']
    annualized=c['income']*12/6
    rate=round(c['total_tax']/c['income']*100,4) if c['income']>0 else 0
    vat_rate=round(c['vat']/c['income']*100,4) if c['income']>0 else 0
    warn_500=rolling_12m>5000000 or annualized>5000000
    wt=''
    if rolling_12m>5000000: wt=f'滚动12月收入超500万(¥{rolling_12m:,.0f})'
    elif annualized>5000000: wt=f'年化收入超500万(¥{annualized:,.0f})'
    small_detail.append({'name':c['name'],'city':c['city'],'income':round(c['income'],2),
                         'vat':round(c['vat'],2),'tax':round(c['total_tax'],2),
                         'rolling_12m':round(rolling_12m,0),'annualized':round(annualized,0),
                         'warn_500':warn_500,'warn_text':wt})

small_by_city={}
for c in small_raw:
    cy=c['city']
    if cy not in small_by_city: small_by_city[cy]={'count':0,'income':0,'tax':0,'warn':0}
    small_by_city[cy]['count']+=1; small_by_city[cy]['income']+=c['income']; small_by_city[cy]['tax']+=c['total_tax']
    d=next((x for x in small_detail if x['name']==c['name']),None)
    if d and d['warn_500']: small_by_city[cy]['warn']+=1

si=sum(c['income'] for c in small_raw); gi=sum(c['income'] for c in general_raw)
wb25.close()

# 预警
ws_warn=wb['增值税税负预警']
alerts={}
for r in range(2, ws_warn.max_row+1):
    sig=ws_warn.cell(r,11).value; n=ws_warn.cell(r,2).value; m=ws_warn.cell(r,1).value
    det=ws_warn.cell(r,12).value; inc=float(ws_warn.cell(r,5).value or 0)
    if sig and sig!='正常' and n and inc>0:
        if n not in alerts: alerts[n]={'months':[],'details':set()}
        alerts[n]['months'].append(m)
        if det and det!='正常': alerts[n]['details'].add(det)

wb.close()
output = {
    'grand_total':gt,
    'cities':cities,
    'companies':companies,
    'monthly':monthly,
    'small_taxpayer':{
        'small_count':len(small_raw),'general_count':len(general_raw),
        'small_income':round(si,2),'general_income':round(gi,2),
        'small_tax':round(sum(c['total_tax'] for c in small_raw),2),
        'general_tax':round(sum(c['total_tax'] for c in general_raw),2),
        'small_rate':round(sum(c['total_tax'] for c in small_raw)/si*100,4) if si>0 else 0,
        'general_rate':round(sum(c['total_tax'] for c in general_raw)/gi*100,4) if gi>0 else 0,
        'by_city':{cy:{'count':v['count'],'income':round(v['income'],2),'tax':round(v['tax'],2),'warn':v['warn']}
                   for cy,v in sorted(small_by_city.items(),key=lambda x:-x[1]['income'])},
        'detail':small_detail
    },
    'alerts':{n:{'count':len(v['months']),'details':list(v['details'])[:3],'months':v['months']}
              for n,v in sorted(alerts.items(),key=lambda x:-len(x[1]['months']))[:20]},
    'months':months_order
}
with open(r'D:\AI分析\税费分析\2026年税费分析看板\data.json','w',encoding='utf-8') as f:
    json.dump(output,f,ensure_ascii=False,indent=2)
print(f"OK: {len(cities)}cities {len(companies)}co {len(small_detail)}small warn={sum(1 for x in small_detail if x['warn_500'])}")
# Auto-build self-contained HTML
import subprocess, os
subprocess.run(['python',os.path.join(os.path.dirname(__file__),'build.py')])
